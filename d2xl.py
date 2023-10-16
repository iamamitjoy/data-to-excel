import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = "Name"
ws.cell(row=1, column=2).value = "Phone Number"
ws.cell(row=1, column=3).value = "Address"
ws.cell(row=1, column=4).value = "Instruction"
ws.cell(row=1, column=5).value = "Reference"


names = []
contacts = []
addresses = []
instructions = []
references = []

i = 0
while True:
    name = input("Enter the NAME of person {}: ".format(i + 1))
    contact = input("Enter the PHONE NUMBER of person {}: ".format(i + 1))
    address = input("Enter the ADDRESS of person {}: ".format(i + 1))
    instruction = input("Enter the INSTRUCTION {}: ".format(i + 1))
    reference = input("Enter the REFERENCE {}: ".format(i + 1))

    if name == "" or contact == "" or address == "" or instruction == "" or reference == "":
        break

    names.append(name)
    contacts.append(contact)
    addresses.append(address)
    instructions.append(instruction)
    references.append(reference)

    i += 1

for i in range(len(names)):
    ws.cell(row=i + 2, column=1).value = names[i]
    ws.cell(row=i + 2, column=2).value = contacts[i]
    ws.cell(row=i + 2, column=3).value = addresses[i]
    ws.cell(row=i + 2, column=4).value = instructions[i]
    ws.cell(row=i + 2, column=5).value = references[i]

wb.save("user_data.xlsx")
